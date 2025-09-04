from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


Headers = ["Item Number", "Description", "BOM", "Template", "Product Type"]


def load_wb_data_only(path: Path) -> object:
    wb = load_workbook(filename=str(path), data_only=True, read_only=False)
    return wb


def find_headers_region(sheet: Worksheet, headers: List[str]) -> Optional[Tuple[int, int]]:
    # Return (row_index, start_col_index) where headers start
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    normalized_headers = [h.strip().lower() for h in headers]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            values: List[str] = []
            for offset, header in enumerate(headers):
                cell = sheet.cell(row=r, column=c + offset)
                values.append(str(cell.value).strip().lower() if cell.value is not None else "")
            if values == normalized_headers:
                return r, c
    return None


def read_generated_items(sheet: Worksheet) -> List[Dict[str, str]]:
    pos = find_headers_region(sheet, Headers)
    if not pos:
        return []
    header_row, start_col = pos
    items: List[Dict[str, str]] = []
    row = header_row + 1
    while True:
        row_vals = [sheet.cell(row=row, column=start_col + i).value for i in range(len(Headers))]
        if all(v in (None, "") for v in row_vals):
            break
        items.append({
            'item_number': str(row_vals[0] or ''),
            'description': str(row_vals[1] or ''),
            'bom': str(row_vals[2] or ''),
            'template': str(row_vals[3] or ''),
            'product_type': str(row_vals[4] or ''),
        })
        row += 1
        if row - header_row > 1000:
            break
    return items


def read_workbook_outputs(path: Path) -> Dict[str, List[Dict[str, str]]]:
    wb = load_wb_data_only(path)
    outputs: Dict[str, List[Dict[str, str]]] = {}
    for name in ["Heater", "Tank", "Pump"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            outputs[name.lower()] = read_generated_items(sheet)
        else:
            outputs[name.lower()] = []
    return outputs


